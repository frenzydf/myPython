import openpyxl
import datetime
import boto3
import enum
from termcolor import colored


class AwsRegion(enum.Enum):
    us_east_1 = 'us-east-1'
    us_west_2 = 'us-west-2'

workbook = openpyxl.load_workbook('whitelist.xlsx')
worksheet = workbook.worksheets[0]
last_row = worksheet.max_row
collect_all_regions = [region.value for region in AwsRegion]

now = datetime.datetime.now()
timestamp = now.strftime("%d%m%Y_%H%M")

def find_addr(address):
    row=1
    for row in range(1,last_row+1):
        cell_value = worksheet['A' + str(row)].value
        if address == cell_value:
            return True, row
    return False, row

def add_addr(rows, ip, id, action):
    worksheet.cell(row=rows,column=1).value = ip
    worksheet.cell(row=rows,column=2).value = timestamp
    worksheet.cell(row=rows,column=3).value = id
    print("{}\t{}\t{}  {}".format(
        colored(rows, 'green').ljust(10),
        colored(ip, 'cyan').ljust(18),
        colored(action, 'green').ljust(18),
        colored(timestamp, 'red')
    ))

print("-"*48)
print("row\tIP\t\tAction\t\t Timestamp")
print("-"*48)
for each_region in collect_all_regions:
    ec2 = boto3.resource('ec2',region_name=each_region)
    for i in ec2.instances.all():
        public_ip = i.public_ip_address
        if public_ip == None:
            pass
        else:
            result = find_addr(public_ip)
            if result[0]:
                add_addr(result[1],public_ip, i.id, "Updated")
            else:
                last_row+=1
                add_addr(last_row,public_ip,i.id,"Added")
    ec2 = boto3.resource('ec2',region_name=each_region)
    for i in ec2.vpc_addresses.all():
        public_ip = i.public_ip
        if public_ip is None:
            continue
        else:
            result = find_addr(public_ip)
            if result[0]:
                add_addr(result[1],public_ip,i.allocation_id,"Updated")
            else:
                last_row+=1
                add_addr(last_row,public_ip,i.allocation_id,"Added")
print("-"*48)
workbook.save('whitelist.xlsx')