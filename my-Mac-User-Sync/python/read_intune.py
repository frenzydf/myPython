import openpyxl

def read_file():
    workbook = openpyxl.load_workbook('../sources/intune.xlsx')
    worksheet = workbook.worksheets[0]
    last_row = worksheet.max_row

    reserved_addresses = []
    hostnames = []
    for i in range(2,last_row+1):
        username = worksheet.cell(row=i,column=1).value
        hostname = worksheet.cell(row=i,column=2).value
        macaddress = worksheet.cell(row=i,column=3).value
        email = worksheet.cell(row=i,column=4).value
        reserved_addres = {
            "hostname": hostname,
            "mac": macaddress,
            "username": username,
            "email": email
        }
        if hostname not in hostnames:
            reserved_addresses.append(reserved_addres)
        hostnames.append(hostname)
    return(reserved_addresses)